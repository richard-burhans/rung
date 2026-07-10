"""Extract dispensary records from a state's discovered list resource.

Dispatches on the coarse list_type stored by state_lists.py:

  pdf    → generic PDF table extraction (pdfplumber)
  csv    → delimited download
  kml    → Google My Maps / KML placemarks
  arcgis → resolve the ArcGIS web map's feature service and query it
  html   → parse the most dispensary-like HTML table
  lookup → search-tool front ends have no static list → AI fallback
  unknown→ AI fallback

Each handler returns DispensaryRecord objects; the caller stamps state + source.
Handlers are best-effort and return [] on failure so the caller can fall back to
the AI extractor (ai_fallback.extract_with_ai), which is only attempted on request.
"""

import asyncio
import csv
import datetime
import html
import io
import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Literal, get_args
from urllib.parse import parse_qs, quote, urljoin, urlparse

import openpyxl
import pdfplumber
import xlrd
from selectolax.parser import HTMLParser

from rung import db
from rung.addresses import (
    PHONE_RE as _PHONE_RE,
    STREET_RE as _STREET_RE,
    ZIP_RE as _ZIP_RE,
    clean as _clean,
    extract_address_blocks as _extract_address_blocks,
    extract_line_blocks as _extract_line_blocks,
)
from rung.http import make_session
from rung.models import DispensaryRecord, LocationObservation
from rung.sources.dedupe import location_key
from rung.text import is_placeholder_name

# Header synonyms, longest/most-specific first so "zip code" beats "code" etc.
_FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "zip_code": ("zip code", "zipcode", "postal code", "zip", "postal"),
    "address": ("street address", "address", "street", "addr"),
    "city": ("city", "town", "municipality"),
    "state": ("state", "province"),
    "phone": ("phone number", "telephone", "phone", "tel"),
    "website": ("website", "web site", "url", "homepage"),
    "name": (
        "dispensary name", "business name", "trade name", "legal name", "dba",
        "licensee", "facility name", "establishment", "entity", "name",
        "company name", "dispensary", "business", "company", "facility", "store",
    ),
}

# Separators states use to pack "NAME – 123 Main St" into one cell.
_NAME_ADDR_SEP_RE = re.compile(r"\s*[–—|]\s*|\s+-\s+")
_YES_NO_VALUES = {"Y", "N", "YES", "NO", "X", "N/A", "NA", "-", "—"}
# iframe srcs that never hold a dispensary list (analytics, embeds, video).
_JUNK_IFRAME_RE = re.compile(
    r"googletagmanager|google-analytics|doubleclick|youtube|vimeo|recaptcha|facebook",
    re.IGNORECASE,
)
# Address primitives (_ZIP_RE, _STREET_RE, _PHONE_RE, _clean,
# _extract_address_blocks) are imported from rung.addresses above;
# that module also serves the company-store extractor, so neither reaches into the other.


def _match_field(header: str) -> str | None:
    """Map a header cell to a DispensaryRecord field, or None."""
    h = " ".join(header.lower().split())
    if not h:
        return None
    for field, synonyms in _FIELD_SYNONYMS.items():
        for syn in synonyms:
            if syn == h or syn in h.split() or h.startswith(syn) or h.endswith(syn):
                return field
    return None


# A bare date — never a street address. The table extractors can mis-map a date column onto
# `address` when a roster interleaves non-dispensary rows (IL's "all cannabis licenses" PDF puts a
# license-issuance date where the address would be); nulling it keeps a junk value out of `address`
# regardless of whether the row's name also tripped the placeholder filter.
_DATE_VALUE_RE = re.compile(r"^\s*\d{1,2}/\d{1,2}/\d{2,4}\s*$")


def _addr_or_none(value: str | None) -> str | None:
    return None if value and _DATE_VALUE_RE.match(value) else value


def _record_from_values(source: str, values: dict[str, str | None]) -> DispensaryRecord:
    """Build a DispensaryRecord from a header→field-name value map.

    The HTML/PDF/CSV extractors map detected columns to DispensaryRecord field names
    (a subset of the text fields — see ``_FIELD_SYNONYMS``), yielding a
    ``dict[str, str | None]``. Pulling each field explicitly (rather than ``**values``)
    keeps the construction type-checkable and silently ignores any non-field key.
    Numeric ``latitude``/``longitude`` are never column-mapped — the geo extractors set
    those directly — so they are intentionally absent here.
    """
    return DispensaryRecord(
        source=source,
        name=values.get("name"),
        address=_addr_or_none(values.get("address")),
        city=values.get("city"),
        state=values.get("state"),
        zip_code=values.get("zip_code"),
        phone=values.get("phone"),
        website=values.get("website"),
    )


# ── HTML tables ──────────────────────────────────────────────────────────────

def _row_cells(tr) -> list[str]:
    return [(c.text() or "").strip() for c in tr.css("td, th")]


def _looks_like_name(value: str) -> bool:
    """A cell that could plausibly be a dispensary name (free text with letters)."""
    v = value.strip()
    if len(v) < 3 or v.upper() in _YES_NO_VALUES:
        return False
    if _ZIP_RE.fullmatch(v) or _PHONE_RE.fullmatch(v):
        return False
    return any(c.isalpha() for c in v)


def _infer_name_column(body_rows: list[list[str]]) -> int | None:
    """Pick the column most likely to hold the dispensary name.

    Used when no header cell matched a name synonym (e.g. the header row is a
    section title like "Southern Nevada Retail Stores"). Choose the column whose
    cells are mostly free text and longest on average — names, often with an
    appended street address, dominate such tables.
    """
    width = max((len(r) for r in body_rows), default=0)
    best_idx: int | None = None
    best_len = 0.0
    for idx in range(width):
        cells = [r[idx] for r in body_rows if idx < len(r)]
        named = [c for c in cells if _looks_like_name(c)]
        if not cells or len(named) < max(3, len(cells) * 0.6):
            continue
        avg_len = sum(len(c) for c in named) / len(named)
        if avg_len > best_len:
            best_len = avg_len
            best_idx = idx
    return best_idx


def _split_name_address(value: str) -> tuple[str, str | None]:
    """Split a combined "NAME – 123 Main St – Adult Use" cell into (name, address).

    Only splits when the segment after the name looks like a street address, so
    names that merely contain a hyphen (e.g. "BEYOND/HELLO - Reno") are left
    intact. A trailing license-type tag (no digits, few words — "Adult Use",
    "Medical") is dropped from the address.
    """
    parts = [p.strip() for p in _NAME_ADDR_SEP_RE.split(value) if p.strip()]
    if len(parts) < 2:
        return value.strip(), None
    name, rest = parts[0], parts[1:]
    if len(rest) > 1 and not any(ch.isdigit() for ch in rest[-1]) and len(rest[-1].split()) <= 4:
        rest = rest[:-1]
    address = ", ".join(rest)
    if name and _STREET_RE.match(address):
        return name, address
    return value.strip(), None


def _location_fraction(records: list[DispensaryRecord]) -> float:
    """Share of records carrying any location signal (address/city/zip/phone)."""
    if not records:
        return 0.0
    located = sum(
        1 for r in records if r.address or r.city or r.zip_code or r.phone
    )
    return located / len(records)


def _street_fraction(cells: list[str]) -> float:
    """Share of the non-empty cells that open with a street number ("226 S Philadelphia Ave")."""
    values = [c.strip() for c in cells if c.strip()]
    if not values:
        return 0.0
    return sum(1 for v in values if _STREET_RE.match(v)) / len(values)


# A header-mapped `address` column holding no street at all is wrong; a lone unmapped column
# that holds streets is where the address really lives. Demand both before overriding a header.
_SWAP_ADDRESS_MAX = 0.10
_SWAP_CANDIDATE_MIN = 0.60


def _repair_swapped_address(col_map: dict[int, str], body: list[list[str]]) -> None:
    """Re-map ``address`` onto the column that actually holds streets. Mutates ``col_map``.

    A roster's header row can be misordered relative to its data. Maryland's dispensary locator
    is headed ``Dispensary | County | Address`` while every data row is ``name | street | county``,
    so trusting the header files the COUNTY as the address and drops the street entirely. The rows
    load and look fine; they simply can never match a company store. Same class of defect as the
    mis-mapped date column `_addr_or_none` nulls, and just as silent.

    Trust the header unless the data contradicts it unambiguously: the mapped `address` column must
    hold essentially no street, and EXACTLY ONE otherwise-unmapped column must mostly hold streets.
    Two candidates, or none, means we cannot tell — so leave the header's word alone. `STREET_RE`
    requires digits *followed by a space*, so a bare licence number ("231001") is not a candidate.
    """
    address_idx = next((i for i, f in col_map.items() if f == "address"), None)
    if address_idx is None or len(body) < 3:
        return

    def column(idx: int) -> list[str]:
        return [row[idx] for row in body if idx < len(row)]

    if _street_fraction(column(address_idx)) > _SWAP_ADDRESS_MAX:
        return
    width = max((len(row) for row in body), default=0)
    candidates = [
        idx for idx in range(width)
        if idx not in col_map and _street_fraction(column(idx)) >= _SWAP_CANDIDATE_MIN
    ]
    if len(candidates) != 1:
        return
    del col_map[address_idx]
    col_map[candidates[0]] = "address"


def _extract_table(table) -> tuple[list[DispensaryRecord], bool]:
    """Extract records from one table.

    Returns (records, header_named). header_named is True when a header cell
    matched a name synonym (high confidence); False when the name column was
    inferred from the body and the caller should require a location signal
    before trusting the table.
    """
    rows = table.css("tr")
    if len(rows) < 2:
        return [], False
    header = _row_cells(rows[0])
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header):
        field = _match_field(cell)
        if field is not None and field not in col_map.values():
            col_map[idx] = field

    header_named = "name" in col_map.values()
    body = rows[1:]
    body_cells = [_row_cells(r) for r in body]
    if not header_named:
        name_idx = _infer_name_column(body_cells)
        if name_idx is None:
            return [], False
        col_map[name_idx] = "name"

    # The header can be misordered relative to the data (MD files its county as the address).
    _repair_swapped_address(col_map, body_cells)

    # If no address column was identified, the name cell may carry the address.
    split_name = "address" not in col_map.values()
    records: list[DispensaryRecord] = []
    for tr in body:
        cells = _row_cells(tr)
        if not cells:
            continue
        values = {field: _clean(cells[idx]) for idx, field in col_map.items() if idx < len(cells)}
        name = values.get("name")
        if not name:
            continue
        if split_name:
            name, address = _split_name_address(name)
            values["name"] = name
            if address:
                values["address"] = address
        records.append(_record_from_values("html", values))
    return records, header_named


def _extract_html(html: str) -> list[DispensaryRecord]:
    """Return dispensary records aggregated across every qualifying HTML table.

    A page may split locations over several tables (e.g. NV's northern/southern
    regions). Header-named tables are trusted outright; inferred-name tables are
    kept only when most rows carry a location signal, to avoid scooping up an
    unrelated text table. Records are de-duplicated on (name, address).
    """
    tree = HTMLParser(html)
    aggregated: list[DispensaryRecord] = []
    seen: set[tuple[str | None, ...]] = set()
    for table in tree.css("table"):
        records, header_named = _extract_table(table)
        if not records:
            continue
        if not header_named and _location_fraction(records) < 0.5:
            continue
        for record in records:
            # Dedup on the full identity, not just (name, address): operators
            # with many locations often share a name and have no street address
            # (e.g. a "Licensee's Name" + City listing), so a coarse key would
            # collapse distinct stores into one.
            key = (record.name, record.address, record.city, record.phone)
            if key not in seen:
                seen.add(key)
                aggregated.append(record)
    return aggregated


# ── PDF tables ───────────────────────────────────────────────────────────────

def _header_map(row: list) -> dict[int, str] | None:
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(row):
        if not cell:
            continue
        field = _match_field(str(cell))
        if field is not None and field not in col_map.values():
            col_map[idx] = field
    return col_map if "name" in col_map.values() else None


def _unmerge_name_overflow(
    name: str | None, address: str | None, city: str | None,
) -> tuple[str | None, str | None]:
    """Repair a name whose overflow was *overprinted* onto the address column.

    PA's roster PDF draws a long store name past its column, on top of the address, at
    overlapping x-positions and in the same font — so pdfplumber interleaves the two runs
    character by character and the name is left truncated::

        name    'Restore Integrative Wellness Center - Elkins'   (truncated: lost 'Park')
        address 'P8a0rk03 Old York Road'                         ('Park' ⊕ '8003')

    No extractor can split that geometrically. But the corruption is *invertible*, because we
    already know one of the two interleaved strings: the lost characters are exactly the tail of
    the ``city``, whose head the truncated ``name`` still ends with. Subtract that tail from the
    address, in order, and both fields come back ('Restore … - Elkins Park', '8003 Old York Road').

    Deliberately strict — every condition must hold, or the row is returned untouched: the name
    must end with a *proper* prefix of the city, every overflow character must be consumed in order
    from the address's head, the repaired address must start with a digit (a street number), and
    both repaired fields must have balanced parentheses. That last guard is not hypothetical: PA
    also carries ``'Ethos - Pittsburgh North of Harmarville (Har'`` / ``'m5a rAvlipllhea) Drive
    East'``, where the overflow came from the name's own parenthetical rather than the city. Its
    real city ("Pittsburgh") already fails the prefix test, but a city of "Harmarville" would have
    slipped through into ``'5 Alpha) Drive East'`` — a stray ``)`` betrays the bad split. That row
    stays broken, which is correct: repairing a row we misread is worse than leaving one broken.
    """
    if not (name and address and city):
        return name, address
    flat_city = city.strip()
    stripped = name.rstrip()
    # The name's tail must be a proper prefix of the city ("… - Elkins" of "Elkins Park").
    overlap = next(
        (n for n in range(len(flat_city) - 1, 0, -1) if stripped.endswith(flat_city[:n])),
        0,
    )
    overflow = flat_city[overlap:].replace(" ", "")
    if not overlap or not overflow:
        return name, address

    kept: list[str] = []
    pending = list(overflow)
    for char in address:
        if pending and char == pending[0]:
            pending.pop(0)
        else:
            kept.append(char)
    repaired = "".join(kept).strip()
    repaired_name = f"{stripped}{flat_city[overlap:]}"
    balanced = all(text.count("(") == text.count(")") for text in (repaired, repaired_name))
    if pending or not repaired[:1].isdigit() or not balanced:
        return name, address  # not the overprint we know how to invert — leave it alone
    return repaired_name, repaired


def _extract_pdf(content: bytes) -> list[DispensaryRecord]:
    records: list[DispensaryRecord] = []
    col_map: dict[int, str] | None = None
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if col_map is None:
                    col_map = _header_map(row)
                    continue  # this row was the header
                # Skip a header repeated atop a later page. A real header matches
                # several field synonyms; require ≥3 so a data row whose name cell
                # merely contains a synonym word ("... Dispensary", "... Store")
                # isn't mistaken for a header and dropped.
                repeated = _header_map(row)
                if repeated is not None and len(repeated) >= 3:
                    continue
                values = {
                    field: _clean(row[idx])
                    for idx, field in col_map.items()
                    if idx < len(row)
                }
                if not values.get("name"):
                    continue
                # A long name overprinted onto the address column interleaves the two (PA).
                values["name"], values["address"] = _unmerge_name_overflow(
                    values.get("name"), values.get("address"), values.get("city"),
                )
                records.append(_record_from_values("pdf", values))
    return records


# ── Arizona DHS 'Licensed Marijuana Establishments' PDF (list_type='az_dhs') ──
# A borderless, word-wrapped columnar PDF the generic pdfplumber table extractor mangles
# (it recovered 25/143 rows and picked the legal entity over the brand). Parse it by
# bucketing words into the fixed columns by x-position instead.

_AZ_CERT_RE = re.compile(r"^0*\d+ES[0-9A-Z]+$")  # e.g. 00000070ESCO78837103
# (column name, left x of its header) — a word belongs to the right-most column it clears.
_AZ_COLS = (
    ("status", 42), ("cert", 116), ("estname", 258), ("dba", 420),
    ("street", 515), ("city", 619), ("zip", 694),
)
_AZ_SKIP_RE = re.compile(
    r"certificate number|establishment name|zip code|street address"
    r"|total licensees|updated|adult use marijuana|licensed marijuana|page",
    re.IGNORECASE,
)


def _az_column(x0: float) -> str:
    column = _AZ_COLS[0][0]
    for name, start in _AZ_COLS:
        if x0 >= start - 6:
            column = name
    return column


def _extract_az_dhs(content: bytes) -> list[DispensaryRecord]:
    """Parse the AZDHS Licensed Marijuana Establishments PDF.

    Words are bucketed into the Status/Cert/Establishment/DBA/Street/City/Zip columns by
    x-position; a record begins on the line carrying a certificate number and absorbs the
    wrapped continuation lines below it (header/footer lines are skipped). The name prefers
    the DBA (storefront brand) and falls back to the legal entity when no DBA is listed.
    """
    fields = ("estname", "dba", "street", "city", "zip")
    rows: list[dict[str, str]] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            lines: dict[int, list] = {}
            for word in page.extract_words():
                lines.setdefault(round(word["top"] / 3), []).append(word)
            current: dict[str, str] | None = None
            for key in sorted(lines):
                cells: dict[str, list[str]] = {}
                for word in sorted(lines[key], key=lambda w: w["x0"]):
                    cells.setdefault(_az_column(word["x0"]), []).append(word["text"])
                joined = {col: " ".join(parts) for col, parts in cells.items()}
                line_text = " ".join(joined.values())
                if _AZ_SKIP_RE.search(line_text):
                    continue
                if _AZ_CERT_RE.match(joined.get("cert", "").replace(" ", "")):
                    current = {col: joined.get(col, "") for col in fields}
                    rows.append(current)
                elif current is not None:  # wrapped continuation of the current record
                    for col in fields:
                        if joined.get(col):
                            current[col] = f"{current[col]} {joined[col]}".strip()
    records: list[DispensaryRecord] = []
    for row in rows:
        name = _clean(row["dba"]) or _clean(row["estname"])
        if name:
            records.append(DispensaryRecord(
                source="az_dhs", name=name, address=_clean(row["street"]),
                city=_clean(row["city"]), zip_code=_clean(row["zip"]),
            ))
    return records


# ── CSV ──────────────────────────────────────────────────────────────────────

_DBA_HEADERS = frozenset({"dba", "d/b/a", "doing business as", "trade name", "trade_name"})


def _extract_csv(text: str) -> list[DispensaryRecord]:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        return []
    # The storefront brand (a DBA / trade-name column) is preferred over the legal/business name
    # via a per-row override below. So the DBA column must NOT claim the `name` slot itself —
    # otherwise a row with a blank DBA but a populated legal name is dropped when the DBA column
    # happens to sort before the legal-name column (the override can't recover it). Keep `name`
    # mapped to a non-DBA column so a blank DBA falls back to the legal/business name.
    dba_idx = next(
        (idx for idx, cell in enumerate(rows[0]) if " ".join(cell.lower().split()) in _DBA_HEADERS),
        None,
    )
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(rows[0]):
        field = _match_field(cell)
        if field is None or field in col_map.values():  # first column per field wins
            continue
        if field == "name" and idx == dba_idx:
            continue  # the DBA column is the override, not the fallback name source
        col_map[idx] = field
    if "name" not in col_map.values():
        # The only name-ish column IS the DBA column (no separate legal/business name) — use it.
        if dba_idx is None:
            return []
        col_map[dba_idx] = "name"
    records: list[DispensaryRecord] = []
    for row in rows[1:]:
        values = {field: _clean(row[idx]) for idx, field in col_map.items() if idx < len(row)}
        if dba_idx is not None and dba_idx < len(row) and _clean(row[dba_idx]):
            values["name"] = _clean(row[dba_idx])
        if values.get("name"):
            records.append(_record_from_values("csv", values))
    return records


# ── Colorado MED 'Stores' Google Sheet CSV (list_type='co_med') ──────────────
# The CO MED publishes its licensed stores as a Google Sheet (exported as CSV). Columns:
# License Number, Facility Name (legal entity), DBA (storefront brand), Facility Type,
# Street, City, ZIP Code, ... The generic CSV path would pick "Facility Name" (legal) for
# the name because it precedes "DBA"; this handler prefers the DBA brand (like az_dhs).

def _extract_co_med(text: str) -> list[DispensaryRecord]:
    reader = csv.DictReader(io.StringIO(text))
    records: list[DispensaryRecord] = []
    for row in reader:
        lower = {(k or "").strip().lower(): v for k, v in row.items()}
        name = _clean(lower.get("dba")) or _clean(lower.get("facility name"))
        if not name:
            continue
        records.append(DispensaryRecord(
            source="co_med", name=name, address=_clean(lower.get("street")),
            city=_clean(lower.get("city")), zip_code=_clean(lower.get("zip code")),
        ))
    return records


# ── Massachusetts CCC 'commenced operations' CSV (list_type='ma_ccc') ────────
# The MA Cannabis Control Commission publishes a SODA CSV of all establishments that have
# commenced operations — geocoded, with every license type. Keep only the active storefronts
# (Marijuana Retailer + Medical Marijuana Treatment Center); drop cultivators/manufacturers/
# transporters/labs. Names are the legal entity (no DBA column), but the rows carry lat/lng.

_MA_STOREFRONT_TYPES = {"Marijuana Retailer", "Medical Marijuana Treatment Center"}


def _extract_ma_ccc(text: str) -> list[DispensaryRecord]:
    def _f(value: str | None) -> float | None:
        try:
            return float(value) if value not in (None, "") else None
        except (TypeError, ValueError):
            return None

    reader = csv.DictReader(io.StringIO(text))
    records: list[DispensaryRecord] = []
    for row in reader:
        if row.get("LICENSE_TYPE") not in _MA_STOREFRONT_TYPES:
            continue
        if (row.get("LICENSE_STATUS") or "").strip() != "Active":
            continue
        name = _clean(row.get("BUSINESS_NAME"))
        if not name:
            continue
        records.append(DispensaryRecord(
            source="ma_ccc", name=name,
            address=_clean(row.get("ADDRESS_1")) or _clean(row.get("PHYSICAL_ADDRESS_1")),
            city=_clean(row.get("CITY")) or _clean(row.get("PHYSICAL_CITY")),
            zip_code=_clean(row.get("ZIP_CODE")) or _clean(row.get("PHYSICAL_ZIP_CODE")),
            latitude=_f(row.get("latitude")), longitude=_f(row.get("longitude")),
        ))
    return records


# ── KML / Google My Maps ─────────────────────────────────────────────────────

_KML_NS = "{http://www.opengis.net/kml/2.2}"


def _extract_kml(content: bytes) -> list[DispensaryRecord]:
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return []
    records: list[DispensaryRecord] = []
    for placemark in root.iter(f"{_KML_NS}Placemark"):
        name = _clean(placemark.findtext(f"{_KML_NS}name"))
        if not name:
            continue
        data: dict[str, str | None] = {}
        ext = placemark.find(f"{_KML_NS}ExtendedData")
        if ext is not None:
            for d in ext.findall(f"{_KML_NS}Data"):
                key = (d.get("name") or "").lower()
                val = d.find(f"{_KML_NS}value")
                field = _match_field(key)
                if field:
                    data[field] = _clean(val.text if val is not None else None)
        address = data.get("address")
        if not address:
            desc = _clean(placemark.findtext(f"{_KML_NS}description"))
            if desc and _STREET_RE.search(desc):
                address = desc
        latitude, longitude = _kml_point(placemark)
        records.append(
            DispensaryRecord(
                source="kml", name=name, address=address,
                city=data.get("city"), state=data.get("state"),
                zip_code=data.get("zip_code"), phone=data.get("phone"),
                latitude=latitude, longitude=longitude,
            )
        )
    return records


def _kml_point(placemark) -> tuple[float | None, float | None]:
    """A placemark's ``<Point>`` coordinates as (lat, lng), or (None, None).

    KML encodes a point as ``lng,lat[,alt]`` (whitespace-padded in Google My Maps
    exports). A 0,0 placeholder or out-of-range value is dropped."""
    point = placemark.find(f"{_KML_NS}Point")
    if point is None:
        return None, None
    raw = _clean(point.findtext(f"{_KML_NS}coordinates"))
    if not raw:
        return None, None
    parts = raw.split(",")
    if len(parts) < 2:
        return None, None
    try:
        longitude, latitude = float(parts[0]), float(parts[1])
    except ValueError:
        return None, None
    if not (-90 <= latitude <= 90) or not (-180 <= longitude <= 180):
        return None, None
    if latitude == 0 and longitude == 0:
        return None, None
    return latitude, longitude


# ── ArcGIS ───────────────────────────────────────────────────────────────────

def _arcgis_attr(attrs: dict, *names: str) -> str | None:
    """Pick the first attribute whose key loosely matches any of names."""
    lower = {k.lower(): v for k, v in attrs.items()}
    for want in names:
        for key, val in lower.items():
            if want in key:
                return _clean(str(val)) if val not in (None, "") else None
    return None


_ARCGIS_SERVICE_RE = re.compile(r"/(?:Feature|Map)Server(?:/\d+)?/?$", re.IGNORECASE)


# ArcGIS feature services cap a single response (the layer's maxRecordCount, often 1000-2000),
# so a layer with more rows MUST be paged with resultOffset or it silently truncates. We page
# until a short page or the server stops flagging exceededTransferLimit; the page cap is a
# runaway guard for a server that ignores resultOffset (it would otherwise re-serve page 0).
_ARCGIS_PAGE_SIZE = 2000
_ARCGIS_PAGE_CAP = 25  # ≤50k features — far above any state's dispensary roster


def _arcgis_float(attrs: dict, *names: str) -> float | None:
    """A numeric attribute picked like :func:`_arcgis_attr`; None when absent/non-numeric."""
    raw = _arcgis_attr(attrs, *names)
    if raw is None:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _arcgis_record(feat: dict) -> DispensaryRecord | None:
    """One ArcGIS feature → a DispensaryRecord, or None when it carries no usable name."""
    attrs = feat.get("attributes") or {}
    name = _arcgis_attr(
        attrs, "dispensar", "name", "dba", "business", "licensee", "facility", "store"
    )
    if not name:
        return None
    # Roster layers often carry lat/lng as plain attributes (Ontario's AGCO does, 100%
    # filled); a coordinate outside the valid range (or a 0,0 placeholder) is dropped.
    # Publishers misspell the field: DC's Open Data layer ships `LONGITDUE`. Match the misspelling
    # too — otherwise the roster loses its geo key, and `compare.py` can only fall back to the
    # address key, which a zip-less roster (DC's) cannot satisfy either. Correct spelling wins:
    # `_arcgis_attr` tries the names in order.
    latitude = _arcgis_float(attrs, "latitude")
    longitude = _arcgis_float(attrs, "longitude", "longitd")
    if (
        latitude is None or longitude is None
        or not (-90 <= latitude <= 90) or not (-180 <= longitude <= 180)
        or (latitude == 0 and longitude == 0)
    ):
        latitude = longitude = None
    return DispensaryRecord(
        source="arcgis", name=name,
        address=_arcgis_attr(attrs, "address", "addr", "street"),
        city=_arcgis_attr(attrs, "city", "town"),
        state=_arcgis_attr(attrs, "state", "province"),
        zip_code=_arcgis_attr(attrs, "zip", "postal"),
        phone=_arcgis_attr(attrs, "phone", "tel"),
        website=_arcgis_attr(attrs, "website", "web", "url"),
        latitude=latitude, longitude=longitude,
    )


async def _query_arcgis_layer(layer_url: str, session) -> list[DispensaryRecord]:
    """Query one feature/map service layer and parse dispensary records, paging all rows.

    A service-root URL (…/FeatureServer) defaults to layer 0. A `where=` query
    string on the URL filters the layer (e.g. to retailer-only license types);
    otherwise all rows are returned. Pages with resultOffset so a layer larger than its
    maxRecordCount isn't silently truncated at the first page.
    """
    parsed = urlparse(layer_url)
    where = parse_qs(parsed.query).get("where", ["1=1"])[0] if parsed.query else "1=1"
    base = parsed._replace(query="").geturl().rstrip("/")
    if not re.search(r"/\d+$", base):
        base = f"{base}/0"
    records: list[DispensaryRecord] = []
    for page in range(_ARCGIS_PAGE_CAP):
        try:
            q = (
                f"{base}/query?where={quote(where)}&outFields=*&f=json"
                f"&resultRecordCount={_ARCGIS_PAGE_SIZE}"
                f"&resultOffset={page * _ARCGIS_PAGE_SIZE}&returnGeometry=false"
            )
            payload = (await session.get(q, timeout=30)).json()
        except Exception:
            break
        features = payload.get("features") or []
        records.extend(rec for feat in features if (rec := _arcgis_record(feat)) is not None)
        # A short page is the end; otherwise keep going only while the server flags more.
        if len(features) < _ARCGIS_PAGE_SIZE or not payload.get("exceededTransferLimit"):
            break
        if page == _ARCGIS_PAGE_CAP - 1:
            print(f"  arcgis: hit page cap ({_ARCGIS_PAGE_CAP}) — layer may be truncated")
    return records


# ── Ontario AGCO cannabis-store map (list_type='on_agco') ────────────────────

# A 32-hex ArcGIS item id anywhere in the configured URL.
_ARCGIS_ITEM_ID_RE = re.compile(r"\b[0-9a-f]{32}\b")


async def _extract_on_agco(url: str, session) -> list[DispensaryRecord]:
    """Extract Ontario's roster from AGCO's ArcGIS Experience app.

    AGCO republishes its layers under date-stamped service names
    (``Authorized_to_open_20250620``), so a pinned FeatureServer URL rotates dead on
    every update; the stable entry point is the Experience-app item. Resolve the app's
    data sources at runtime and query the "Authorized To Open" layer — its siblings
    (Application_in_progress / Cancelled_Authorizations / Public_Notice) are other
    lifecycle stages, not the open-store roster.
    """
    match = _ARCGIS_ITEM_ID_RE.search(url)
    if not match:
        return []
    try:
        data = (
            await session.get(
                f"https://www.arcgis.com/sharing/rest/content/items/{match.group(0)}/data?f=json",
                timeout=30,
            )
        ).json()
    except Exception:
        return []
    for source in (data.get("dataSources") or {}).values():
        if not isinstance(source, dict):
            continue
        for child in (source.get("childDataSourceJsons") or {}).values():
            layer_url = child.get("url") or ""
            if "authorized_to_open" in layer_url.lower():
                return await _query_arcgis_layer(layer_url, session)
    return []


# ── Alberta AGLC cannabis-licensee report (list_type='ab_aglc') ──────────────

# The report's stable header names (verified live 2026-07-04); parsed by name so a
# column reorder doesn't silently shift fields.
_AGLC_FIELDS = {
    "name": "Establishment Name",
    "city": "Site City Name",
    "address": "Site Address Line 1",
    "address2": "Site Address Line 2",
    "province": "Site Province Abbrev",
    "zip_code": "Site Postal Code",
    "phone": "Telephone Number",
}


def _xls_text(cell) -> str | None:
    """A cell's value as clean text. Numeric cells (phone numbers, license ids) come
    back as floats from xlrd — fold 2508842774.0 → '2508842774'."""
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return str(int(cell.value))
    return _clean(str(cell.value))


def _aglc_record(values: dict[str, str | None]) -> DispensaryRecord | None:
    """One AGLC report row → a record, keeping only Alberta retail sites.

    The licensee report also lists out-of-province supplier/producer sites (ON/BC/QC…,
    152 of 955 rows on the 2026-07-04 pull) — those are not Alberta stores, and the
    persist path would otherwise stamp them state=AB.
    """
    if values.get("province") != "AB":
        return None
    name = values.get("name")
    if not name:
        return None
    address = values.get("address")
    if values.get("address2"):
        address = f"{address}, {values['address2']}" if address else values["address2"]
    return DispensaryRecord(
        source="ab_aglc", name=name, address=address,
        city=values.get("city"), state="AB",
        zip_code=values.get("zip_code"), phone=values.get("phone"),
    )


def _extract_ab_aglc(content: bytes) -> list[DispensaryRecord]:
    """Extract Alberta's roster from AGLC's licensee report (legacy OLE2 .xls).

    Direct download (no auth): aglc.ca/cannabis/cannabis-licensee-report/EXCEL.
    Header row 0, one licensee per row; no website or coordinate columns exist
    (menu handles and coords come from Stage 2 / the pool bootstrap).
    """
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheets()[0]
    if sheet.nrows < 2:
        return []
    header = [_clean(str(sheet.cell_value(0, c))) for c in range(sheet.ncols)]
    columns = {field: header.index(col) for field, col in _AGLC_FIELDS.items() if col in header}
    if "name" not in columns:
        return []
    records = []
    for r in range(1, sheet.nrows):
        values = {field: _xls_text(sheet.cell(r, c)) for field, c in columns.items()}
        record = _aglc_record(values)
        if record is not None:
            records.append(record)
    return records


# ── Saskatchewan SLGA cannabis retailers (list_type='sk_slga') ───────────────

# The SLGA retailers spreadsheet's stable header names (verified live 2026-07-04).
_SLGA_FIELDS = {
    "name": "Operating Name",
    "city": "City",
    "address": "Street Address",
    "website": "Website",
    "status": "StatusDesc",
}
# The download link on the SLGA page is date-stamped (…-excel--june-30.xlsx) and rotates,
# so resolve the current retailers .xlsx href from the page rather than pinning it.
_SLGA_XLSX_RE = re.compile(r'href="([^"]*cannabis-retailers-excel[^"]*\.xlsx)"', re.IGNORECASE)


def _slga_record(values: dict[str, str | None]) -> DispensaryRecord | None:
    """One SLGA spreadsheet row → a record, keeping only active retailers."""
    if (values.get("status") or "").strip().lower() != "active":
        return None
    name = values.get("name")
    if not name:
        return None
    website = values.get("website")
    if website and website.strip().upper() in ("N/A", "NA", "-"):
        website = None
    return DispensaryRecord(
        source="sk_slga", name=name, address=values.get("address"),
        city=values.get("city"), state="SK", website=website,
    )


def _extract_slga_xlsx(content: bytes) -> list[DispensaryRecord]:
    """Parse the SLGA cannabis-retailers .xlsx (header row 0, one retailer per row)."""
    import io

    book = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = book.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    header = [_clean(str(c)) if c is not None else None for c in next(rows, [])]
    columns = {field: header.index(col) for field, col in _SLGA_FIELDS.items() if col in header}
    if "name" not in columns:
        return []
    records = []
    for row in rows:
        values = {
            field: (_clean(str(row[i])) if i < len(row) and row[i] is not None else None)
            for field, i in columns.items()
        }
        record = _slga_record(values)
        if record is not None:
            records.append(record)
    return records


async def _extract_sk_slga(url: str, session) -> list[DispensaryRecord]:
    """Extract Saskatchewan's roster from SLGA's authorized-retailers page.

    Resolves the date-stamped retailers .xlsx link from the page at runtime (the URL
    rotates on each monthly update), downloads it, and parses the active retailers.
    """
    try:
        page = (await session.get(url, timeout=30)).text
    except Exception:
        return []
    match = _SLGA_XLSX_RE.search(page)
    if not match:
        return []
    xlsx_url = urljoin(url, match.group(1))
    try:
        resp = await session.get(xlsx_url, timeout=60)
    except Exception:
        return []
    if resp.status_code >= 400 or resp.content[:4] != b"PK\x03\x04":
        return []
    return _extract_slga_xlsx(resp.content)


# ── British Columbia LCRB establishments map (list_type='bc_lcrb') ────────────

def _bc_lcrb_record(obj: dict) -> DispensaryRecord | None:
    """One LCRB establishment object → a record.

    The open JSON API behind justice.gov.bc.ca/lcrb/map (the app is open-source,
    bcgov/jag-lcrb-carla-public): name/addressStreet/addressCity/addressPostal/
    latitude/longitude/phone/isOpen. Licensed-but-not-yet-open rows (isOpen=false)
    are kept, mirroring Ontario's authorized-to-open roster.
    """
    name = _clean(obj.get("name"))
    if not name:
        return None
    return DispensaryRecord(
        source="bc_lcrb", name=name,
        address=_clean(obj.get("addressStreet")),
        city=_clean(obj.get("addressCity")),
        state="BC",
        zip_code=_clean(obj.get("addressPostal")),
        phone=_clean(obj.get("phone")),
        latitude=obj.get("latitude"), longitude=obj.get("longitude"),
    )


# Virginia's CCA dispensary page is a Squarespace site: no `<table>`, and the addresses are not in
# the visible DOM as text we can block-parse. Each dispensary is a map block whose JSON config is
# stashed in an HTML *attribute* (so it arrives entity-escaped, `&quot;addressLine1&quot;`). Unescape
# the document, then lift each `"location": {...}` object. The objects are flat (no nested braces),
# which is what makes the non-greedy `[^{}]*` safe here rather than needing a real parser.
_VA_LOCATION_RE = re.compile(r'"location":(\{[^{}]*\})')
# addressLine2 is always "City, VA, ZIP" on this page (verified: 23/23, 2026-07-09).
_VA_LINE2_RE = re.compile(r"^\s*(?P<city>[^,]+),\s*(?P<state>[A-Z]{2}),?\s*(?P<zip>\d{5})")


def _va_cca_record(loc: dict) -> DispensaryRecord | None:
    """One CCA map-block `location` object → a DispensaryRecord, or None if it isn't a dispensary."""
    name = _clean(str(loc.get("addressTitle") or ""))
    street = _clean(str(loc.get("addressLine1") or ""))
    if not name or not street:
        return None                      # a non-dispensary map block (the page carries one)
    city = state = zip_code = None
    if match := _VA_LINE2_RE.match(str(loc.get("addressLine2") or "")):
        city, state, zip_code = (match["city"].strip(), match["state"], match["zip"])
    latitude, longitude = loc.get("markerLat"), loc.get("markerLng")
    if not (isinstance(latitude, (int, float)) and isinstance(longitude, (int, float))
            and -90 <= latitude <= 90 and -180 <= longitude <= 180
            and not (latitude == 0 and longitude == 0)):
        latitude = longitude = None
    return DispensaryRecord(
        source="va_cca", name=name, address=street, city=city, state=state,
        zip_code=zip_code, phone=None, website=None,
        latitude=latitude, longitude=longitude,
    )


# ── Atlist (my.atlist.com embedded maps) ─────────────────────────────────────
# New Jersey's CRC "Find a Dispensary" page renders an Atlist map in an <iframe>. The page's own
# HTML holds ONE table, and it lists DELIVERY SERVICES — parsing it yielded 3 rows named
# "Passing Puff Delivery", "TerpTaxi" and "Weedies" while we held 186 NJ company stores. (The
# neighbouring `/dispensaries/roll-up/` page is a product-RECALL table, not a roster. Both are
# traps for a table-hunting parser.)
#
# The real roster is the map's markers. Atlist's SPA reads `/v1/map/{id}/markers` and, for a
# publicly-shared map, sends the literal bearer token `public` — its bundle falls back to the
# string "public" when no user token exists. That is the anonymous read path the embed itself uses.
_ATLIST_HOST = "https://api.atlist.com"
_ATLIST_MAP_ID_RE = re.compile(r"/map/([0-9a-fA-F-]{36})")
# "460 Maple Ave, Elizabeth, NJ 07202, USA" — Google-formatted, so the shape is stable.
_ATLIST_ADDRESS_RE = re.compile(
    r"^(?P<street>.+?),\s*(?P<city>[^,]+),\s*(?P<state>[A-Z]{2})\s+(?P<zip>\d{5})(?:-\d{4})?"
    r"(?:,\s*[A-Za-z .]+)?$"
)
_ATLIST_MAX_PAGES = 40   # 250 markers/page; a runaway-cursor backstop, not an expected limit


def _atlist_record(marker: dict) -> DispensaryRecord | None:
    """One Atlist marker → a roster record. Coordinates are kept even when the address won't parse.

    A marker whose `formattedAddress` is a road rather than a street number ("NJ-66, Neptune
    Township, NJ, USA") still carries lat/long, and `compare`'s proximity tier pairs on those alone.
    Dropping it because the text is unparseable would lose a real licensee.
    """
    name = _clean(marker.get("name"))
    if not name:
        return None
    latitude, longitude = _coord(marker.get("lat")), _coord(marker.get("long"))
    parsed = _ATLIST_ADDRESS_RE.match((marker.get("formattedAddress") or "").strip())
    if parsed is None:
        return DispensaryRecord(source="atlist", name=name, website=_clean(marker.get("buttonLink")),
                                latitude=latitude, longitude=longitude)
    return DispensaryRecord(
        source="atlist", name=name, address=_clean(parsed["street"]), city=_clean(parsed["city"]),
        state=parsed["state"], zip_code=parsed["zip"], website=_clean(marker.get("buttonLink")),
        latitude=latitude, longitude=longitude,
    )


def _coord(value: object) -> float | None:
    if isinstance(value, int | float):
        return float(value)
    return None


async def _extract_atlist(url: str, session) -> list[DispensaryRecord]:
    """Every marker on a publicly-shared Atlist map, paging through its opaque `nextToken`."""
    found = _ATLIST_MAP_ID_RE.search(url)
    if not found:
        return []
    endpoint = f"{_ATLIST_HOST}/v1/map/{found.group(1)}/markers"
    headers = {"Authorization": "Bearer public"}
    records: list[DispensaryRecord] = []
    token: str | None = None
    for _page in range(_ATLIST_MAX_PAGES):
        params = {"nextToken": token} if token else None
        try:
            response = await session.get(endpoint, headers=headers, params=params, timeout=60)
        except Exception:
            return records
        if response.status_code >= 400:
            return records
        try:
            payload = response.json()
        except ValueError:
            return records
        if not isinstance(payload, dict):
            return records
        for marker in payload.get("markers") or []:
            if isinstance(marker, dict):
                record = _atlist_record(marker)
                if record is not None:
                    records.append(record)
        token = payload.get("nextToken")
        if not token:
            break
    return records


async def _extract_va_cca(url: str, session) -> list[DispensaryRecord]:
    """Virginia CCA's medical-cannabis dispensary locations (Squarespace map blocks)."""
    try:
        response = await session.get(url, timeout=60, allow_redirects=True)
    except Exception:
        return []
    if response.status_code >= 400:
        return []
    document = html.unescape(response.text or "")
    records: list[DispensaryRecord] = []
    seen: set[tuple[str, str]] = set()
    for match in _VA_LOCATION_RE.finditer(document):
        try:
            location = json.loads(match.group(1))
        except ValueError:
            continue
        if not isinstance(location, dict):
            continue
        record = _va_cca_record(location)
        if record is None:
            continue
        identity = ((record.name or "").lower(), (record.address or "").lower())
        if identity in seen:                 # the same block can render twice (desktop + mobile)
            continue
        seen.add(identity)
        records.append(record)
    return records


async def _extract_bc_lcrb(url: str, session) -> list[DispensaryRecord]:
    """Fetch BC's roster from the LCRB map API (a bare JSON list, no paging/auth)."""
    try:
        payload = (await session.get(url, timeout=60)).json()
    except Exception:
        return []
    if not isinstance(payload, list):
        return []
    return [rec for obj in payload if isinstance(obj, dict)
            if (rec := _bc_lcrb_record(obj)) is not None]


async def _extract_arcgis(url: str, session) -> list[DispensaryRecord]:
    """Extract from an ArcGIS source.

    Accepts either a direct feature/map service URL (queried straight away) or a
    webappviewer/Experience app URL (with an ?id= item), which is resolved to its
    operational layers — directly or via a nested web map.
    """
    # A direct feature/map service URL — query it straight away.
    if _ARCGIS_SERVICE_RE.search(urlparse(url).path):
        return await _query_arcgis_layer(url, session)

    item_id = None
    for part in urlparse(url).query.split("&"):
        if part.startswith("id="):
            item_id = part[3:]
            break
    if not item_id:
        return []
    sharing = url.split("/apps/", 1)[0] + "/sharing/rest"

    async def _item_data(iid: str) -> dict:
        try:
            return (
                await session.get(
                    f"{sharing}/content/items/{iid}/data?f=json", timeout=30
                )
            ).json()
        except Exception:
            return {}

    def _layers(d: dict) -> list[str]:
        return [
            layer["url"]
            for layer in (d.get("operationalLayers") or [])
            if isinstance(layer, dict) and layer.get("url")
        ]

    # The viewer's id is the *app* item. If its data already lists operational
    # layers, use them; a Web AppBuilder app instead nests the web map id under
    # map.itemId, whose data carries the operational layers.
    data = await _item_data(item_id)
    layer_urls = _layers(data)
    if not layer_urls:
        webmap_id = (
            (data.get("map") or {}).get("itemId")
            or (data.get("values") or {}).get("webmap")
        )
        if webmap_id:
            layer_urls = _layers(await _item_data(webmap_id))

    for layer_url in layer_urls:
        records = await _query_arcgis_layer(layer_url, session)
        if records:
            return records
    return []


# ── California DCC retailer API (list_type='ca_dcc') ─────────────────────────

# California bounding box (minLat, maxLat, minLng, maxLng).
_CA_DCC_BBOX = (32.3, 42.1, -124.6, -114.0)
_CA_DCC_PAGE = 1000  # the API caps a response at this many rows and ignores ?page


def _ca_dcc_record(lic: dict) -> DispensaryRecord | None:
    """Map one DCC license object to a record, keeping only active retailers."""
    if lic.get("licenseStatus") != "Active":
        return None
    if "retailer" not in (lic.get("licenseType") or "").lower():
        return None
    name = _clean(lic.get("businessDbaName") or lic.get("businessLegalName"))
    if not name:
        return None
    return DispensaryRecord(
        source="ca_dcc", name=name,
        address=_clean(lic.get("premiseStreetAddress")),
        city=_clean(lic.get("premiseCity")),
        state=_clean(lic.get("premiseState")),
        zip_code=_clean(lic.get("premiseZipCode")),
        phone=_clean(lic.get("businessPhone")),
        latitude=lic.get("premiseLatitude"),
        longitude=lic.get("premiseLongitude"),
    )


async def _extract_ca_dcc(url: str, session) -> list[DispensaryRecord]:
    """Sweep California's DCC RetailerLocationSearch API for active retailers.

    The endpoint returns {metadata:{totalCount,…}, data:[license, …]} for a lat/long
    box but caps a response at 1000 rows and ignores ?page. So we recursively split
    any box whose totalCount exceeds one page into quadrants until every box fits,
    then dedupe distinct premises.
    """
    records: list[DispensaryRecord] = []
    seen: set[tuple[str | None, str | None]] = set()

    async def sweep(box: tuple[float, float, float, float], depth: int) -> None:
        min_lat, max_lat, min_lng, max_lng = box
        q = (
            f"{url}?minLatitude={min_lat}&maxLatitude={max_lat}"
            f"&minLongitude={min_lng}&maxLongitude={max_lng}&pageSize={_CA_DCC_PAGE}"
        )
        try:
            payload = (await session.get(q, timeout=60)).json()
        except Exception:
            return
        data = payload.get("data") or []
        total = (payload.get("metadata") or {}).get("totalCount", len(data))
        if total > len(data) and depth < 6:  # capped — subdivide into quadrants
            mid_lat, mid_lng = (min_lat + max_lat) / 2, (min_lng + max_lng) / 2
            await sweep((min_lat, mid_lat, min_lng, mid_lng), depth + 1)
            await sweep((min_lat, mid_lat, mid_lng, max_lng), depth + 1)
            await sweep((mid_lat, max_lat, min_lng, mid_lng), depth + 1)
            await sweep((mid_lat, max_lat, mid_lng, max_lng), depth + 1)
            return
        for lic in data:
            record = _ca_dcc_record(lic)
            if record is None:
                continue
            key = (record.name, record.address)
            if key not in seen:
                seen.add(key)
                records.append(record)

    await sweep(_CA_DCC_BBOX, 0)
    return records


# ── Dispatcher ───────────────────────────────────────────────────────────────

# The list_type vocabulary this dispatcher understands. `state_lists._classify` (the
# producer) must stay a subset of this — guarded by test_list_type_vocabulary_consistent
# so the two never drift. Any other value falls through to the html parser.
ListType = Literal[
    "pdf", "csv", "kml", "arcgis", "atlist", "ca_dcc", "az_dhs", "co_med", "ma_ccc",
    "on_agco", "ab_aglc", "bc_lcrb", "sk_slga", "va_cca", "lookup", "html",
]
HANDLED_LIST_TYPES: frozenset[str] = frozenset(get_args(ListType))


async def extract_records(list_url: str, list_type: ListType | str) -> list[DispensaryRecord]:
    """Extract dispensary records from a list resource. Returns [] on failure.

    `list_type` should be one of `HANDLED_LIST_TYPES`; an unrecognised value falls through
    to the html table/address-block parser (the same path as `"html"`).
    """
    async with make_session() as session:
        if list_type == "arcgis":
            return await _extract_arcgis(list_url, session)
        if list_type == "ca_dcc":
            return await _extract_ca_dcc(list_url, session)
        if list_type == "on_agco":
            return await _extract_on_agco(list_url, session)
        if list_type == "bc_lcrb":
            return await _extract_bc_lcrb(list_url, session)
        if list_type == "va_cca":
            return await _extract_va_cca(list_url, session)
        if list_type == "atlist":
            return await _extract_atlist(list_url, session)
        if list_type == "sk_slga":
            return await _extract_sk_slga(list_url, session)
        if list_type in ("lookup",):
            return []  # dynamic search front ends — caller uses AI fallback
        try:
            resp = await session.get(list_url, timeout=60, allow_redirects=True)
        except Exception:
            return []
        if resp.status_code >= 400:
            return []

        try:
            # A URL ending in .pdf sometimes serves an HTML error/redirect page;
            # trust the bytes, not the extension.
            if list_type == "pdf" and resp.content[:5].startswith(b"%PDF"):
                return _extract_pdf(resp.content)
            if list_type == "az_dhs" and resp.content[:5].startswith(b"%PDF"):
                return _extract_az_dhs(resp.content)
            # OLE2 magic — an .xls URL sometimes serves an HTML error page; trust the bytes.
            if list_type == "ab_aglc" and resp.content[:4] == b"\xd0\xcf\x11\xe0":
                return _extract_ab_aglc(resp.content)
            if list_type == "co_med":
                return _extract_co_med(resp.text)
            if list_type == "ma_ccc":
                return _extract_ma_ccc(resp.text)
            if list_type == "csv":
                return _extract_csv(resp.text)
            if list_type == "kml":
                return _extract_kml(resp.content)
            # html / unknown, or a mislabeled pdf that was really HTML: try the
            # table extractor, then the non-table card/list fallback.
            return _extract_page(resp.text)
        except Exception:
            return []


# ── Browser-rendered HTML (opt-in) ───────────────────────────────────────────

def _content_iframes(html: str) -> list[str]:
    """Return http(s) iframe srcs that might embed a dispensary list/app."""
    tree = HTMLParser(html)
    urls: list[str] = []
    for frame in tree.css("iframe"):
        src = frame.attributes.get("src") or ""
        if src.startswith("http") and not _JUNK_IFRAME_RE.search(src):
            urls.append(src)
    return urls


def _extract_page(html: str) -> list[DispensaryRecord]:
    """Tables, then the address-block fallback, then line blocks.

    Strictly ordered, each a LAST RESORT for the one before, so a page that already yields
    rows cannot change behaviour: the line-block rung only ever runs where we currently get
    **nothing**. It reads a street line followed by a "City, ST zip" line — the shape
    `BLOCK_ADDRESS_RE` cannot span, because that needs a comma between street and city.
    Alabama's AMCC roster is prose of exactly this form.
    """
    return _extract_html(html) or _extract_address_blocks(html) or _extract_line_blocks(html)


async def extract_rendered(url: str, tab) -> list[DispensaryRecord]:
    """Render a JS-driven page in Chrome and extract its dispensaries.

    Tries the table extractor then the card/list extractor on the rendered DOM,
    falling back to the largest content iframe when the top document yields
    nothing (covers Tableau/ArcGIS/embedded-locator iframes).
    """
    from rung.browser import render_html

    html = await render_html(tab, url)
    if not html:
        return []
    records = _extract_page(html)
    if records:
        return records
    for iframe_url in _content_iframes(html)[:2]:
        records = _extract_page(await render_html(tab, iframe_url))
        if records:
            return records
    return []


async def _render_empty_targets(
    targets: list,
    extracted: dict[str, list[DispensaryRecord]],
    methods: dict[str, str],
) -> None:
    """Render each still-empty target in one shared Chrome session, in place."""
    from pydoll.browser.chromium import Chrome

    from rung.browser import make_browser_options

    async with Chrome(options=make_browser_options()) as browser:
        tab = await browser.start()
        for rec in targets:
            try:
                records = await extract_rendered(rec.list_url, tab)
            except Exception:
                records = []
            if records:
                extracted[rec.abbr] = records
                methods[rec.abbr] = "render"


# ── Orchestration ────────────────────────────────────────────────────────────

@dataclass
class ExtractResult:
    abbr: str
    name: str
    list_type: str | None
    count: int
    method: str  # 'static' | 'render' | 'ai' | 'none'


def record_roster_observations(
    conn: db.DBConn, state: str, records: list[DispensaryRecord],
    *, now: datetime.datetime | None = None,
) -> int:
    """Append store-lifecycle history from a state's just-extracted roster. Caller commits.

    The Stage-1 ``state_roster`` twin of the overlay's ``company_site`` capture: build one
    :class:`LocationObservation` per physical location (``dedupe.geo_key``, ``address_key``
    fallback — the roster is only partially geocoded, so the address key carries more weight
    here) and hand them to the shared engine (``db.record_location_observations``). Runs only
    on a NON-EMPTY extraction — a failed roster fetch records nothing, so a store's observed
    absence stays a real signal rather than an artifact of a dead list URL. Roster rows carry
    no menu handle; the state-registered name goes in ``operator`` (raw, canonicalized at
    read). See docs/store_history_design.md. Returns observation rows appended.
    """
    observations: dict[str, LocationObservation] = {}
    for record in records:
        key = location_key(record.latitude, record.longitude, record.address, record.zip_code)
        if not key or key in observations:
            continue  # unidentifiable (e.g. MD's county-only rows), or a duplicate rooftop row
        observations[key] = LocationObservation(
            location_key=key, state=state, latitude=record.latitude,
            longitude=record.longitude, address=record.address, city=record.city,
            zip_code=record.zip_code, operator=record.name,
        )
    return db.record_location_observations(conn, "state_roster", observations, now=now)


async def run_extract_states(
    conn: db.DBConn,
    only: set[str] | None = None,
    use_ai: bool = False,
    use_render: bool = False,
    record_history: bool = False,
) -> list[ExtractResult]:
    """Extract dispensary records for every state with a discovered list URL.

    Three tiers run in order of cost: static handlers concurrently; then, when
    use_render is set, one shared Chrome renders JS-driven pages that came back
    empty; then, when use_ai is set, the local-Ollama fallback. Each state's
    existing rows are replaced atomically and only when extraction yields data,
    so re-runs are idempotent and a dead URL never wipes prior good rows.

    ``record_history`` (opt-in): also append store-lifecycle history
    (``state_roster`` observations) alongside each non-empty state replace, in the
    same commit — see :func:`record_roster_observations`.
    """
    from rung.db import (
        delete_dispensaries_for_state,
        get_all_state_programs,
        insert_dispensary,
    )

    targets = [
        r for r in get_all_state_programs(conn)
        if r.list_url and r.list_status in ("found", "override", "stored")
        and (only is None or r.abbr in only)
    ]

    # Tier 1 — static extraction, concurrently.
    async def _extract(rec):
        records = await extract_records(rec.list_url, rec.list_type or "html")
        return rec, records

    raw = await asyncio.gather(*(_extract(r) for r in targets))
    extracted: dict[str, list[DispensaryRecord]] = {rec.abbr: recs for rec, recs in raw}
    methods: dict[str, str] = {
        rec.abbr: ("static" if recs else "none") for rec, recs in raw
    }

    # Tier 2 — browser render of still-empty JS pages (one shared Chrome).
    if use_render:
        renderable = [
            r for r in targets
            if not extracted[r.abbr] and (r.list_type in (None, "html", "unknown"))
        ]
        if renderable:
            await _render_empty_targets(renderable, extracted, methods)

    # Tier 3 — AI fallback (per state) + persist.
    results: list[ExtractResult] = []
    for rec in targets:
        records = extracted[rec.abbr]
        method = methods[rec.abbr]
        if not records and use_ai and rec.list_url:
            try:
                from rung.sources.ai_fallback import extract_with_ai
                records = await extract_with_ai(rec.list_url, source_tag="ai")
                method = "ai" if records else "none"
            except Exception as exc:
                # The AI tier is opt-in and best-effort, so one state's failure must not abort
                # the run — but surface it (Ollama down, bad import, schema error) so a
                # misconfigured tier isn't silently indistinguishable from a genuinely empty page.
                print(f"  ai fallback failed for {rec.abbr}: {exc}")
                method = "none"

        # Replace this state's rows only when we actually extracted something.
        # Drop non-operator junk rows (license-number/header/test stubs) so they never enter
        # `dispensaries`. Filter BEFORE the guard below, so a junk-only roster counts as 0 real
        # records and keeps the prior good rows rather than wiping them.
        records = [record for record in records if not is_placeholder_name(record.name)]
        # A transient failure or a list URL that has gone 404 yields 0 records;
        # wiping the prior good rows in that case would be data loss.
        if records:
            delete_dispensaries_for_state(conn, rec.abbr)
            for record in records:
                record.state = rec.abbr  # scope rows to the state for idempotent replace
                insert_dispensary(conn, record)
            if record_history:
                # Store-lifecycle history from the fresh roster, same commit as the replace.
                record_roster_observations(conn, rec.abbr, records)
            conn.commit()

        results.append(
            ExtractResult(rec.abbr, rec.name, rec.list_type, len(records), method)
        )
    return results


def print_extract_report(results: list[ExtractResult]) -> None:
    rows = sorted(results, key=lambda r: (-r.count, r.name))
    sep = "-" * 70
    print(sep)
    print(f"{'State':<22} | {'Type':<8} | {'Method':<7} | Records")
    print(sep)
    total = 0
    with_data = 0
    for r in rows:
        total += r.count
        if r.count:
            with_data += 1
        print(f"{r.name:<22} | {(r.list_type or '-'):<8} | {r.method:<7} | {r.count}")
    print(sep)
    print(f"States with records: {with_data}/{len(results)} | Total records: {total}")
    print(sep)
